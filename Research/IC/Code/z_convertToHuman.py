import os
import csv
from datetime import datetime
import openpyxl
import base64

CONVERT_2_AND_3_TO_0 = True

SCRIPT_FOLDER = os.path.dirname(os.path.abspath(__file__))
INPUT_FOLDER = os.path.join(SCRIPT_FOLDER, "toConvert")
OUTPUT_FOLDER = os.path.join(SCRIPT_FOLDER, "toConvert/output")

input_files = ["mateus.csv", "evandro.xlsx", "cezio.xlsx", "human_cezio_350.csv", "conciliado_first220.csv"]
output_file = "output.csv"

def decode_date_from_id(file_id):
    try:
        decoded = base64.b64decode(file_id).decode('utf-8')
        # Extract date part (remove .txt extension)
        date_str = decoded.replace('.txt', '').strip()
        # Format: DDMMYYYY -> DD/MM/YYYY
        if len(date_str) == 8 and date_str.isdigit():
            day = date_str[0:2]
            month = date_str[2:4]
            year = date_str[4:8]
            return f"{day}/{month}/{year}"
    except Exception as e:
        print(f"  Warning: Could not decode date from ID '{file_id}': {e}")
    
    # Return current date as fallback
    return datetime.now().strftime("%d/%m/%Y")

def ensure_output_folder():
    if not os.path.exists(OUTPUT_FOLDER):
        os.makedirs(OUTPUT_FOLDER)

def read_csv_file(filepath):
    data = []
    with open(filepath, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        for row in reader:
            if len(row) >= 3:
                file_id = row[0].strip()
                phrase = row[1].strip()
                sentiment_str = row[2].strip()
                
                # Ignore if sentiment is empty
                if not sentiment_str:
                    continue
                
                try:
                    sentiment = int(sentiment_str)
                    data.append((file_id, phrase, sentiment))
                except ValueError:
                    # Skip rows with invalid sentiment values
                    print(f"  Warning: Skipping row with invalid sentiment: '{sentiment_str}'")
                    continue
    return data

def read_csv_file_pipe_delimited(filepath):
    """Read CSV file with pipe delimiter (format: cod|texto|cezio)"""
    data = []
    with open(filepath, 'r', encoding='utf-8') as f:
        for line_num, line in enumerate(f, 1):
            # Skip header line if it contains 'cod|texto|cezio'
            if line_num == 1 and 'cod|texto|cezio' in line.lower():
                continue
            
            line = line.strip()
            if not line:
                continue
            
            parts = line.split('|')
            if len(parts) >= 3:
                file_id = parts[0].strip()
                phrase = parts[1].strip()
                sentiment_str = parts[2].strip()
                
                # Ignore if sentiment is empty
                if not sentiment_str:
                    continue
                
                try:
                    sentiment = int(float(sentiment_str))
                    data.append((file_id, phrase, sentiment))
                except ValueError:
                    # Skip rows with invalid sentiment values
                    print(f"  Warning: Skipping row {line_num} with invalid sentiment: '{sentiment_str}'")
                    continue
    return data

def read_csv_file_pipe_delimited2(filepath):
    """Read CSV file with pipe delimiter (format: cod|texto|conciliado)"""
    data = []
    with open(filepath, 'r', encoding='utf-8') as f:
        for line_num, line in enumerate(f, 1):
            # Skip header line if it contains 'cod|texto|conciliado'
            if line_num == 1 and 'cod|texto|conciliado' in line.lower():
                continue
            
            line = line.strip()
            if not line:
                continue
            
            parts = line.split('|')
            if len(parts) >= 3:
                file_id = parts[0].strip()
                phrase = parts[1].strip()
                sentiment_str = parts[2].strip()
                
                # Ignore if sentiment is empty
                if not sentiment_str:
                    continue
                
                try:
                    sentiment = int(float(sentiment_str))
                    data.append((file_id, phrase, sentiment))
                except ValueError:
                    # Skip rows with invalid sentiment values
                    print(f"  Warning: Skipping row {line_num} with invalid sentiment: '{sentiment_str}'")
                    continue
    return data

def read_xlsx_file(filepath):
    data = []
    workbook = openpyxl.load_workbook(filepath)
    sheet = workbook.active
    
    for row in sheet.iter_rows(min_row=1, values_only=True):
        if row[0] is not None and row[1] is not None and row[2] is not None:
            file_id = str(row[0]).strip()
            phrase = str(row[1]).strip()
            sentiment_str = str(row[2]).strip()
            
            # Ignore if sentiment is empty
            if not sentiment_str:
                continue
            
            # Try to convert to int, skip if invalid
            try:
                # Remove '=' if present (Excel formula prefix)
                sentiment_str = sentiment_str.replace('=', '')
                sentiment = int(float(sentiment_str))
                data.append((file_id, phrase, sentiment))
            except (ValueError, TypeError):
                # Skip rows with invalid sentiment values
                print(f"  Warning: Skipping row with invalid sentiment: '{sentiment_str}'")
                continue
    
    workbook.close()
    return data

def read_xlsx_file_alt_format(filepath):
    data = []
    workbook = openpyxl.load_workbook(filepath)
    sheet = workbook.active
    
    for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header row
        if row[0] is not None and row[1] is not None and row[2] is not None:
            file_id = str(row[0]).strip()
            sentiment_str = str(row[1]).strip()
            phrase = str(row[2]).strip()
            
            # Ignore if sentiment is empty
            if not sentiment_str:
                continue
            
            # Try to convert to int, skip if invalid
            try:
                # Remove '=' if present (Excel formula prefix)
                sentiment_str = sentiment_str.replace('=', '')
                sentiment = int(float(sentiment_str))
                data.append((file_id, phrase, sentiment))
            except (ValueError, TypeError):
                # Skip rows with invalid sentiment values
                print(f"  Warning: Skipping row with invalid sentiment: '{sentiment_str}'")
                continue
    
    workbook.close()
    return data

def convert_to_output_format(data):
    output_lines = []
    
    for file_id, phrase, sentiment in data:
        date = decode_date_from_id(file_id)
        output_line = f"{date}|human|{sentiment}|{phrase}"
        output_lines.append(output_line)
    
    return output_lines

def main():
    ensure_output_folder()
    
    all_output_lines = []
    all_output_lines.append("Date|Model|Grade|Sentence")
    
    for filename in input_files:
        filepath = os.path.join(INPUT_FOLDER, filename)
        
        if not os.path.exists(filepath):
            print(f"Warning: File not found: {filepath}")
            continue
        
        print(f"Processing: {filename}")
        
        # Determine file type and read accordingly
        if filename.endswith('.csv'):
            # Check if it's human_cezio_350.csv with pipe delimiter format
            if filename == 'human_cezio_350.csv':
                data = read_csv_file_pipe_delimited(filepath)
            elif filename == 'conciliado_first220.csv':
                data = read_csv_file_pipe_delimited2(filepath)
            else:
                data = read_csv_file(filepath)
        elif filename.endswith('.xlsx'):
            # Check if it's cezio.xlsx with alternative format
            if filename == 'cezio.xlsx':
                data = read_xlsx_file_alt_format(filepath)
            else:
                data = read_xlsx_file(filepath)
        else:
            print(f"Warning: Unsupported file format: {filename}")
            continue
        
        # Convert to output format
        output_lines = convert_to_output_format(data)
        all_output_lines.extend(output_lines)
        
        print(f"  Processed {len(data)} entries from {filename}")
    
    # Convert grades -2 and -3 to 0
    if CONVERT_2_AND_3_TO_0:
        for i in range(1, len(all_output_lines)):  # Skip header
            parts = all_output_lines[i].split('|')
            if len(parts) == 4:
                try:
                    grade = int(parts[2])
                    if grade in [-2, -3]:
                        parts[2] = '0'
                        all_output_lines[i] = '|'.join(parts)
                except ValueError:
                    continue
    
    # Write all output to file
    output_path = os.path.join(OUTPUT_FOLDER, output_file)
    with open(output_path, 'w', encoding='utf-8') as f:
        for line in all_output_lines:
            f.write(line + '\n')
    
    print(f"\nTotal entries processed: {len(all_output_lines)}")
    print(f"Output saved to: {output_path}")

if __name__ == "__main__":
    main()


